import os
from openpyxl import load_workbook


class ReadExcel:

    def __init__(self, fileName="file.xlsx"):
        try:  # TO FIND THE FILE IN THE PREVIOUS LEVEL OR 2 LEVELS BEHIND
            directory = os.path.abspath('.')
        except:
            directory = os.path.abspath('..')

        filepathRead = directory + "\\Files\\"  # NAVIGATE TO THE 'Files' FOLDER WHICH CONTAINS THE file.xlsx
        filepathRead = filepathRead + fileName  # COMBINE THE FILE PATH WITH THE FILE NAME
        self.wb = load_workbook(filepathRead)  # LOAD THE EXCEL FILE AND STORE IT IN THE wb OBJECT

    # RETURNS ALL CELLS IN A GIVEN COLUMN
    def getColumnDataFromSheet(self, sheet="Page1", column=1):
        sheet = self.wb[sheet]
        columnData = []
        for i in range(1, sheet.max_row + 1):
            columnData.append(sheet['{0}{1}'.format(column, i)].value)
        return columnData

    # RETURNS ALL CELLS IN A GIVEN ROW
    def getRowDataFromSheet(self, sheet="Page1", row=1):
        sheet = self.wb[sheet]
        rowData = []
        letters = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
        for i in range(0, sheet.max_column):
            rowData.append(sheet['{0}{1}'.format(letters[i], row)].value)
        return rowData

    def getCellFromSheet(self, sheet="Page1", cell="A1"):
        sheet = self.wb[sheet]
        return sheet[str(cell)].value

    def change_color(self, sheet="Setup", cell="A1"):
        sheet = self.wb[sheet]
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import PatternFill, colors
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        sheet[cell].fill = redFill
        self.wb.save(filename="aaa.xlsx")

    def getMaxRow(self, sheet="Page1", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_row

    def getMaxColumn(self, sheet="Page1", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_column
